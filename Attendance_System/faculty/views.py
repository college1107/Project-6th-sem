from django.shortcuts import render
from college_admin.models import *
from User.models import *
from datetime import datetime
from django.http import HttpResponse
from django.views import View
from pymongo import *
from openpyxl import Workbook
from faculty.utils import *
from django.contrib import messages
import time
import xlwt
import pandas as pd

def Add_Attendance_to_postgres(date):
    data = FetchColumn('attendance_system', 'en_no')
    for en in data:
        register_instance = register.objects.get(en_no=en[0])
        attendance_data = register_instance.attended
        row_column(attendance_data,en,date)

def F_home(request):
    data = register.objects.values("en_no", "name", "attended")
    current_time = time.localtime()
    current_datetime = datetime.fromtimestamp(time.mktime(current_time))
    formatted_date = current_datetime.strftime("%Y-%m-%d")
    context = {
            "page": "Faculty",
            "data": data,
            "current_datetime": current_datetime,
        }
    date = ''
    if request.method == "POST":
        date = request.POST.get("date")
        if not date:
            messages.success(request, "Please add Date")
            context.update({"color": "danger"})
            return render(request, "F_index.html", context)


    # ***************************************************************************************
    AddData(data)
    CreateColumn('attendance_system',formatted_date,'BOOLEAN')
    Add_Attendance_to_postgres(formatted_date)
    # ***************************************************************************************
    return render(request, "F_index.html", context)


def download_excel_data(request):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response[
        "Content-Disposition"
    ] = f'attachment; filename="{time.strftime("%d-%m-%Y")}.xlsx"'

    wb = Workbook()
    ws = wb.active

    columns = ["En_no", "Name", "Attended"]

    for col_num, column in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=column)

    queryset = register.objects.all().values("en_no", "name", "attended")

    for row_num, row_data in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=row_data["en_no"])
        ws.cell(row=row_num, column=2, value=row_data["name"])
        ws.cell(row=row_num, column=3, value=row_data["attended"])

    wb.save(response)

    return response
